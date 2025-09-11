# features/spect_viewer/gui/side_panel.py

from pathlib import Path
from typing import Dict, Any, List
from datetime import datetime
from PySide6.QtWidgets import QSizePolicy, QHeaderView, QSplitter

from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFrame, QScrollArea,
    QTableWidget, QTableWidgetItem, QAbstractItemView, QCheckBox, QHeaderView
)
from PySide6.QtGui import QColor, QFont

from .bsi_canvas import BSICanvas
from features.spect_viewer.logic.quantification_integration import QuantificationManager
# CHANGED: Cleaned up imports. BSI_EXPORT_BUTTON_STYLE is no longer needed
# as we will use the unified EXPORT_CHART_BUTTON_STYLE for all main export buttons.
from core.gui.ui_constants import PRIMARY_BUTTON_STYLE

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# --- Helper Class for Clickable Header ---
class ClickableFrame(QFrame):
    """A QFrame that emits a 'clicked' signal when pressed."""
    clicked = Signal()
    def mousePressEvent(self, event):
        self.clicked.emit()
        super().mousePressEvent(event)

# This combined stylesheet is now used for ALL primary export buttons in this panel.
# It uses the primary style and adds a specific state for when it's disabled.
EXPORT_CHART_BUTTON_STYLE = PRIMARY_BUTTON_STYLE + """
    QPushButton:disabled {
        background-color: #cccccc;
        color: #666666;
        border: none;
    }
"""
import logging
class BSISidePanel(QWidget):
    """
    REFACTORED: BSI Side Panel with a collapsible chart section and single view support.
    """
    export_requested = Signal(str)
    analysis_requested = Signal()
    scan_selected = Signal(str)

    def __init__(self, parent: QWidget = None):
        super().__init__(parent)
        self.current_patient_folder = None
        self.current_patient_id = None
        self.current_study_date = None
        self.quant_manager = QuantificationManager()
        self.resize(500, self.height())
        self._build_ui()

    def _build_ui(self):
        main_panel_layout = QVBoxLayout(self)
        main_panel_layout.setContentsMargins(0, 0, 0, 0)
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.NoFrame)
        content_widget = QWidget()
        scroll_area.setWidget(content_widget)
        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(8, 8, 8, 8)
        content_layout.setSpacing(12)

        self._create_title_section(content_layout)
        
        content_layout.addWidget(self._create_collapsible_chart_section())

        content_layout.addWidget(self._create_results_table_v2())
        content_layout.addWidget(self._create_controls_section())
        content_layout.addStretch()
        main_panel_layout.addWidget(scroll_area)
        
    def _create_collapsible_chart_section(self) -> QWidget:
        """Creates the entire collapsible section containing the chart and its controls."""
        section_container = QFrame()
        section_layout = QVBoxLayout(section_container)
        section_layout.setContentsMargins(0, 0, 0, 0)
        section_layout.setSpacing(0)

        header = ClickableFrame()
        header.setCursor(Qt.PointingHandCursor)
        header.setStyleSheet("""
            ClickableFrame {
                background-color: #4A5568; /* gray-700 */
                border: 1px solid #2D3748; /* gray-800 */
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
            }
        """)
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(10, 8, 10, 8)
        
        title = QLabel("<b>BSI Trend Chart & Tools</b>")
        title.setStyleSheet("color: white; font-size: 12px;")
        
        self.chart_toggle_btn = QPushButton("▼")
        self.chart_toggle_btn.setFixedSize(24, 24)
        self.chart_toggle_btn.setStyleSheet("""
            QPushButton {
                color: white; border: none; background-color: transparent;
                font-size: 14px; font-weight: bold;
            }
        """)
        self.chart_toggle_btn.setFocusPolicy(Qt.NoFocus)

        header_layout.addWidget(title)
        header_layout.addStretch()
        header_layout.addWidget(self.chart_toggle_btn)

        self.chart_content_frame = QFrame()
        self.chart_content_frame.setStyleSheet("""
            QFrame {
                border: 1px solid #e2e8f0; /* gray-200 */
                border-top: none;
                border-bottom-left-radius: 6px;
                border-bottom-right-radius: 6px;
                background-color: #ffffff;
            }
        """)
        content_layout = QVBoxLayout(self.chart_content_frame)
        content_layout.setContentsMargins(8, 8, 8, 8)
        content_layout.setSpacing(10)
        
        self.bsi_canvas = BSICanvas()
        content_layout.addWidget(self.bsi_canvas)
        
        content_layout.addWidget(self._create_chart_controls_section())

        export_layout = QHBoxLayout()
        self.export_chart_btn = QPushButton("Export Chart")
        # Apply the unified primary button style
        self.export_chart_btn.setStyleSheet(EXPORT_CHART_BUTTON_STYLE)
        self.export_chart_btn.clicked.connect(lambda: self.export_requested.emit("chart"))
        export_layout.addStretch()
        export_layout.addWidget(self.export_chart_btn)
        content_layout.addLayout(export_layout)

        section_layout.addWidget(header)
        section_layout.addWidget(self.chart_content_frame)

        header.clicked.connect(self._toggle_chart_visibility)
        self.chart_toggle_btn.clicked.connect(self._toggle_chart_visibility)
        
        return section_container

    def _toggle_chart_visibility(self):
        """Toggles the visibility of the chart content frame."""
        is_visible = self.chart_content_frame.isVisible()
        self.chart_content_frame.setVisible(not is_visible)
        self.chart_toggle_btn.setText("▲" if not is_visible else "▼")

    def _on_chart_visibility_changed(self):
        """Handle chart visibility checkbox changes and update export button state."""
        anterior_visible = self.anterior_checkbox.isChecked()
        posterior_visible = self.posterior_checkbox.isChecked()
        
        self.bsi_canvas.set_line_visibility(anterior_visible, posterior_visible)
        
        self._update_export_chart_button_state()

    def _update_export_chart_button_state(self):
        """
        Updates the enabled/disabled state of the Export Chart button based on
        data availability AND checkbox state.
        """
        if not hasattr(self, 'export_chart_btn'):
            return

        has_data = self.current_patient_id is not None
        is_any_line_visible = self.anterior_checkbox.isChecked() or self.posterior_checkbox.isChecked()
        
        self.export_chart_btn.setEnabled(has_data and is_any_line_visible)

    def _create_title_section(self, layout):
        title_frame = QFrame()
        title_frame.setStyleSheet("background: #f8f9fa; border: 1px solid #e9ecef; border-radius: 6px; padding: 8px;")
        title_layout = QVBoxLayout(title_frame)
        self.title_label = QLabel("<b>BSI Quantification Analysis </b>")
        self.title_label.setStyleSheet("font-size: 14px; color: #2c3e50; font-weight: bold; margin-bottom: 4px;")
        title_layout.addWidget(self.title_label)
        layout.addWidget(title_frame)
        
    def _create_chart_controls_section(self) -> QWidget:
        """Create checkbox controls for BSI chart lines and returns the widget."""
        controls_frame = QFrame()
        controls_frame.setStyleSheet("background: #f0f8ff; border: 1px solid #cce7ff; border-radius: 4px; padding: 8px; margin: 4px 0px;")
        controls_layout = QVBoxLayout(controls_frame)
        
        controls_title = QLabel("<b>Chart Display Options</b>")
        controls_title.setStyleSheet("font-size: 11px; color: #495057; font-weight: bold; margin-bottom: 4px;")
        controls_layout.addWidget(controls_title)
        
        checkboxes_layout = QHBoxLayout()
        
        self.anterior_checkbox = QCheckBox("Anterior")
        self.anterior_checkbox.setChecked(True)
        self.anterior_checkbox.setStyleSheet("color: #ff6b6b; font-weight: bold;")
        self.anterior_checkbox.stateChanged.connect(self._on_chart_visibility_changed)
        checkboxes_layout.addWidget(self.anterior_checkbox)
        
        self.posterior_checkbox = QCheckBox("Posterior")
        self.posterior_checkbox.setChecked(True)
        self.posterior_checkbox.setStyleSheet("color: #4ecdc4; font-weight: bold;")
        self.posterior_checkbox.stateChanged.connect(self._on_chart_visibility_changed)
        checkboxes_layout.addWidget(self.posterior_checkbox)
        
        checkboxes_layout.addStretch()
        controls_layout.addLayout(checkboxes_layout)
        return controls_frame
        
    def _create_controls_section(self) -> QWidget:
        """Creates the section for exporting table data."""
        controls_frame = QFrame()
        controls_frame.setStyleSheet("background: #f8f9fa; border: 1px solid #e9ecef; border-radius: 4px; padding: 8px;")
        controls_layout = QVBoxLayout(controls_frame)
        
        buttons_layout = QHBoxLayout()
        
        self.export_excel_btn = QPushButton("Export Excel")
        self.export_excel_btn.setStyleSheet(EXPORT_CHART_BUTTON_STYLE) 
        self.export_excel_btn.clicked.connect(self._export_excel_data)  # ADD
        buttons_layout.addWidget(self.export_excel_btn)
        
        buttons_layout.addStretch()
        controls_layout.addLayout(buttons_layout)
        return controls_frame
        
    def _update_button_states(self, has_data: bool):
        if hasattr(self, 'export_excel_btn'):
            self.export_excel_btn.setEnabled(has_data)
        
        self._update_export_chart_button_state()

    def _wire_two_tables_once(self):
        if getattr(self, "_two_tables_wired", False): return
        right_v, left_v = self.results_table_right.verticalScrollBar(), self.results_table_left.verticalScrollBar()
        right_v.valueChanged.connect(left_v.setValue)
        left_v.valueChanged.connect(right_v.setValue)
        self.results_table_left.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.results_table_right.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        right_vh, left_vh = self.results_table_right.verticalHeader(), self.results_table_left.verticalHeader()
        def _on_row_height_changed(logical_index: int, old_size: int, new_size: int):
            if logical_index < self.results_table_left.rowCount(): left_vh.resizeSection(logical_index, new_size)
        right_vh.sectionResized.connect(_on_row_height_changed)
        def _full_resync_heights():
            rows = min(self.results_table_left.rowCount(), self.results_table_right.rowCount())
            for r in range(rows): left_vh.resizeSection(r, self.results_table_right.rowHeight(r))
        model = self.results_table_right.model()
        model.rowsInserted.connect(lambda *args: _full_resync_heights())
        model.rowsRemoved.connect(lambda *args: _full_resync_heights())
        self._two_tables_wired = True

    def _create_results_table_v2(self) -> QWidget:
        table_container = QFrame()
        table_container.setObjectName("bsiTableFrame")
        table_container.setStyleSheet("""
            QFrame#bsiTableFrame { border: 1px solid #e9ecef; border-radius: 6px; background: #ffffff; }
            QTableWidget { border: none; }
            QHeaderView::section { background: #f8f9fa; border: none; border-bottom: 1px solid #e9ecef; padding: 6px; }
        """)
        table_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        outer = QVBoxLayout(table_container)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)
        title = QLabel("<b>Quantification Results</b>")
        title.setStyleSheet("font-size: 12px; color: #495057; font-weight: bold; margin: 8px;")
        outer.addWidget(title)

        # 🟢 PENTING: Gunakan QSplitter sebagai container untuk kedua tabel
        splitter = QSplitter(Qt.Horizontal)
        splitter.setHandleWidth(1) # Atur lebar handle agar terlihat seperti garis tipis
        splitter.setStyleSheet("""
            QSplitter::handle {
                background-color: #e9ecef;
                width: 1px;
                margin: 0px;
            }
        """)

        # LEFT (Region, frozen)
        self.results_table_left = QTableWidget()
        self.results_table_left.setColumnCount(1)
        self.results_table_left.setHorizontalHeaderLabels(["Region"])
        self.results_table_left.verticalHeader().setVisible(False)
        self.results_table_left.setAlternatingRowColors(True)
        self.results_table_left.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.results_table_left.setSelectionMode(QAbstractItemView.NoSelection)
        self.results_table_left.setSortingEnabled(False)
        self.results_table_left.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.results_table_left.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        
        # ❗ PENTING: Set stretch last section to False for the left table
        self.results_table_left.horizontalHeader().setStretchLastSection(False)
        self.results_table_left.verticalHeader().setDefaultSectionSize(30)
        self.results_table_left.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
        
        # RIGHT (data)
        self.results_table_right = QTableWidget()
        self.results_table_right.setColumnCount(4)
        self.results_table_right.setHorizontalHeaderLabels([
            "Malignant Ant", "Malignant Post", "Benign Ant", "Benign Post"
        ])
        self.results_table_right.verticalHeader().setVisible(False)
        self.results_table_right.setAlternatingRowColors(True)
        self.results_table_right.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.results_table_right.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.results_table_right.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.results_table_right.setSortingEnabled(False)
        header = self.results_table_right.horizontalHeader()
        header.setSortIndicatorShown(False)
        header.setSectionsClickable(False)
        
        # ❗ PENTING: Set stretch last section to True for the right table
        header.setStretchLastSection(True)
        header.setSectionResizeMode(QHeaderView.Interactive)
        self.results_table_right.verticalHeader().setDefaultSectionSize(30)
        self.results_table_right.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.results_table_right.setMinimumHeight(650)
        
        # 🟢 PENTING: Tambahkan kedua tabel ke splitter
        splitter.addWidget(self.results_table_left)
        splitter.addWidget(self.results_table_right)
        
        # Atur rasio lebar awal
        splitter.setSizes([200, 600]) 

        # 🟢 PENTING: Tambahkan splitter ke layout luar
        outer.addWidget(splitter) 

        note = QLabel("<i>Format: pixel_count (decimal_ratio) • Drag header untuk resize</i>")
        note.setStyleSheet("font-size: 10px; color: #6c757d; font-style: italic; margin: 6px 8px;")
        outer.addWidget(note)
        self._wire_two_tables_once()
        return table_container

    def _sync_left_order_with_right(self):
        order = []
        rows = self.results_table_right.rowCount()
        for r in range(rows):
            it0 = self.results_table_right.item(r, 0)
            if not it0 or it0.data(Qt.UserRole) == "zzz_total": continue
            order.append((r, it0.data(Qt.UserRole + 1)))
        for dst_row, (_, region_key) in enumerate(order):
            self.results_table_left.setItem(dst_row, 0, QTableWidgetItem((region_key or "").title()))
        last = len(order)
        if last < self.results_table_left.rowCount():
            total_item = QTableWidgetItem("TOTAL")
            f = QFont(); f.setBold(True); total_item.setFont(f); total_item.setData(Qt.UserRole, "zzz_total")
            self.results_table_left.setItem(last, 0, total_item)
    
    def load_patient_data(self, patient_folder: Path, patient_id: str, study_date: str) -> bool:
        try:
            self.current_patient_folder, self.current_patient_id, self.current_study_date = patient_folder, patient_id, study_date
            canvas_patient_folder = patient_folder.parent if len(patient_folder.name) == 8 and patient_folder.name.isdigit() else patient_folder
            self.bsi_canvas.load_bsi_data(canvas_patient_folder, patient_id, study_date)
            quant_results = self.quant_manager.load_quantification_results(patient_folder, patient_id, study_date)
            if quant_results:
                anterior_data = quant_results.get('anterior_results', {}).get('bsi_results', {}) if quant_results.get('anterior_results') else {}
                posterior_data = quant_results.get('posterior_results', {}).get('bsi_results', {}) if quant_results.get('posterior_results') else {}
                processing_mode = quant_results.get('summary_statistics', {}).get('processing_mode', 'unknown')
                self._populate_results_table_v2(anterior_data, posterior_data, processing_mode)
                self._update_patient_info(patient_id, study_date)
                self._update_button_states(True)
            else:
                if hasattr(self, 'results_table_left'): self.results_table_left.setRowCount(0)
                if hasattr(self, 'results_table_right'): self.results_table_right.setRowCount(0)
                self._update_patient_info(patient_id, "No Data Found")
                self._update_button_states(False)
            return True
        except Exception:
            import traceback; traceback.print_exc()
            self.clear_patient_data()
            return False
            
    def _populate_results_table_v2(self, anterior_data: dict, posterior_data: dict, processing_mode: str = "unknown"):
        if not anterior_data and not posterior_data:
            self.results_table_left.setRowCount(0); self.results_table_right.setRowCount(0)
            return

        all_regions = sorted(list(set(anterior_data.keys()) | set(posterior_data.keys()) - {'background'}))
        rows = len(all_regions) + 1
        self.results_table_left.setRowCount(rows); self.results_table_right.setRowCount(rows)

        total_ant_benign, total_post_benign, total_ant_malignant, total_post_malignant = 0, 0, 0, 0

        for row, region_name in enumerate(all_regions):
            ant = anterior_data.get(region_name, {}) or {}; pos = posterior_data.get(region_name, {}) or {}
            
            ant_benign, ant_benign_ratio = ant.get('benign_pixels', 0), ant.get('benign_ratio', 0.0)
            pos_benign, pos_benign_ratio = pos.get('benign_pixels', 0), pos.get('benign_ratio', 0.0)
            ant_malig, ant_malig_ratio = ant.get('malignant_pixels', 0), ant.get('malignant_ratio', 0.0)
            pos_malig, pos_malig_ratio = pos.get('malignant_pixels', 0), pos.get('malignant_ratio', 0.0)

            total_ant_benign += ant_benign; total_post_benign += pos_benign
            total_ant_malignant += ant_malig; total_post_malignant += pos_malig

            if processing_mode == 'single_view_anterior':
                pos_benign_text, pos_malig_text = "N/A", "N/A"
            else:
                pos_benign_text = f"{pos_benign} ({pos_benign_ratio:.3f})"
                pos_malig_text = f"{pos_malig} ({pos_malig_ratio:.3f})"

            if processing_mode == 'single_view_posterior':
                ant_benign_text, ant_malig_text = "N/A", "N/A"
            else:
                ant_benign_text = f"{ant_benign} ({ant_benign_ratio:.3f})"
                ant_malig_text = f"{ant_malig} ({ant_malig_ratio:.3f})"

            # LEFT (Region)
            self.results_table_left.setItem(row, 0, QTableWidgetItem(region_name.title()))

            # RIGHT (4 kolom)
            right_items = [
                QTableWidgetItem(ant_malig_text), QTableWidgetItem(pos_malig_text),
                QTableWidgetItem(ant_benign_text), QTableWidgetItem(pos_benign_text)
            ]
            for c, it in enumerate(right_items):
                it.setData(Qt.UserRole + 1, region_name)
                if c in (0, 1):
                    if c == 0 and ant_malig > 0 and processing_mode != 'single_view_posterior':
                        it.setBackground(QColor(255, 200, 200))
                    elif c == 1 and pos_malig > 0 and processing_mode != 'single_view_anterior':
                        it.setBackground(QColor(255, 200, 200))
                    elif it.text() == "N/A":
                        it.setBackground(QColor(240, 240, 240))
                self.results_table_right.setItem(row, c, it)

        # TOTAL ROW
        last = len(all_regions)
        total_lbl = QTableWidgetItem("TOTAL"); f = QFont(); f.setBold(True); total_lbl.setFont(f); total_lbl.setData(Qt.UserRole, "zzz_total")
        self.results_table_left.setItem(last, 0, total_lbl)

        if processing_mode == 'single_view_anterior':
            totals = [QTableWidgetItem(str(total_ant_malignant)), QTableWidgetItem("N/A"),
                      QTableWidgetItem(str(total_ant_benign)), QTableWidgetItem("N/A")]
        elif processing_mode == 'single_view_posterior':
            totals = [QTableWidgetItem("N/A"), QTableWidgetItem(str(total_post_malignant)),
                      QTableWidgetItem("N/A"), QTableWidgetItem(str(total_post_benign))]
        else:
            totals = [QTableWidgetItem(str(total_ant_malignant)), QTableWidgetItem(str(total_post_malignant)),
                      QTableWidgetItem(str(total_ant_benign)), QTableWidgetItem(str(total_post_benign))]

        for it in totals:
            it.setBackground(QColor(230, 230, 230))
            it.setData(Qt.UserRole, "zzz_total")
        for c, it in enumerate(totals):
            it.setData(Qt.UserRole + 1, "__total__")
            self.results_table_right.setItem(last, c, it)
        
        # PENTING: Gunakan header kanan untuk mengatur lebar awal kolom
        self.results_table_right.resizeColumnsToContents()
        self.results_table_right.setColumnWidth(0, max(130, self.results_table_right.columnWidth(0)))
        self.results_table_right.setColumnWidth(1, max(130, self.results_table_right.columnWidth(1)))
        self.results_table_right.setColumnWidth(2, max(120, self.results_table_right.columnWidth(2)))
        self.results_table_right.setColumnWidth(3, max(120, self.results_table_right.columnWidth(3)))

        # Hitung lebar tabel kiri secara dinamis & atur sebagai lebar tetap
        left_table = self.results_table_left
        left_header = left_table.horizontalHeader()
        left_header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        left_table.resizeColumnsToContents()
        width = left_table.verticalHeader().width() + left_header.sectionSize(0) + left_table.frameWidth() * 2
        left_table.setFixedWidth(width)
        left_header.setSectionResizeMode(0, QHeaderView.Interactive)

        # Samakan tinggi semua baris awal
        rows_sync = min(self.results_table_left.rowCount(), self.results_table_right.rowCount())
        for r in range(rows_sync):
            self.results_table_left.verticalHeader().resizeSection(r, self.results_table_right.rowHeight(r))

    def _update_patient_info(self, patient_id: str, study_date: str):
        pass
    
    def clear_patient_data(self):
        logging.info("[BSI SIDE PANEL] Clearing data...")
        self.current_patient_folder, self.current_patient_id, self.current_study_date = None, None, None
        if hasattr(self, 'bsi_canvas'): self.bsi_canvas.clear_data()
        if hasattr(self, 'results_table_left'): self.results_table_left.setRowCount(0)
        if hasattr(self, 'results_table_right'): self.results_table_right.setRowCount(0)
        self._update_patient_info("N/A", "N/A")
        self._update_button_states(False)

    def set_session_code(self, session_code: str): self._current_session_code = session_code
    
    def refresh_current_patient(self):
        if self.current_patient_folder and self.current_patient_id:
            logging.info(f"[BSI PANEL] Refreshing for {self.current_patient_id}")
            if self.load_patient_data(self.current_patient_folder, self.current_patient_id, self.current_study_date or "latest"): logging.info("[BSI PANEL] Refresh successful")
            else: logging.info("[BSI PANEL] Refresh failed")
        else: logging.info("[BSI PANEL] No patient to refresh")
    
    def _export_excel_data(self):
        if not self.current_patient_id or not self.current_study_date:
            return
            
        try:
            from PySide6.QtWidgets import QFileDialog
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
            
            filename = f"BSI_Results_{self.current_patient_id}_{self.current_study_date}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Export BSI Excel", filename, "Excel Files (*.xlsx)"
            )
            if not file_path:
                return
                
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "BSI Results"
            
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Title
            ws['A1'] = f"BSI Quantification Results - {self.current_patient_id}"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:E1')
            
            # Headers (row 3)
            headers = ["Region", "Malignant Ant", "Malignant Post", "Benign Ant", "Benign Post"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for r in range(self.results_table_left.rowCount()):
                row_num = r + 4
                
                # Region name
                region_cell = ws.cell(row=row_num, column=1)
                region_cell.value = self.results_table_left.item(r, 0).text() if self.results_table_left.item(r, 0) else ""
                region_cell.border = border
                
                # Data columns
                for c in range(4):
                    data_cell = ws.cell(row=row_num, column=c + 2)
                    data_cell.value = self.results_table_right.item(r, c).text() if self.results_table_right.item(r, c) else ""
                    data_cell.border = border
                    data_cell.alignment = Alignment(horizontal='center')
                    
                    # Highlight malignant cells (red background)
                    if c in (0, 1) and "N/A" not in str(data_cell.value):
                        try:
                            pixel_count = int(data_cell.value.split()[0])
                            if pixel_count > 0:
                                data_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                        except:
                            pass
            
            if hasattr(self, 'quant_manager') and self.quant_manager.current_results:
                summary = self.quant_manager.current_results.get('summary_statistics', {})
                processing_mode = summary.get('processing_mode', 'unknown')
                
                bsi_row_num = ws.max_row + 1
                
                # Left cell: "BSI Score"
                bsi_label_cell = ws.cell(row=bsi_row_num, column=1, value="BSI Score")
                bsi_label_cell.font = Font(bold=True)
                bsi_label_cell.border = border
                bsi_label_cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                
                # Right merged cell: BSI value
                if processing_mode == 'dual_view':
                    bsi_text = f"Anterior: {summary.get('anterior_bsi', 0.0):.2f}% | Posterior: {summary.get('posterior_bsi', 0.0):.2f}%"
                elif processing_mode == 'single_view_anterior':
                    bsi_text = f"Anterior: {summary.get('anterior_bsi', 0.0):.2f}% (Posterior: N/A)"
                elif processing_mode == 'single_view_posterior':
                    bsi_text = f"Posterior: {summary.get('posterior_bsi', 0.0):.2f}% (Anterior: N/A)"
                else:
                    bsi_text = f"BSI: {summary.get('bsi_score', 0.0):.2f}%"
                
                # Merge cells B to E untuk BSI value
                ws.merge_cells(f'B{bsi_row_num}:E{bsi_row_num}')
                bsi_value_cell = ws.cell(row=bsi_row_num, column=2, value=bsi_text)
                bsi_value_cell.font = Font(bold=True)
                bsi_value_cell.border = border
                bsi_value_cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                bsi_value_cell.alignment = Alignment(horizontal='center')
            # Auto-adjust column widths
            # Auto-adjust column widths (skip merged cells)
            column_widths = {
                'A': 15,  # Region
                'B': 12,  # Malignant Ant  
                'C': 12,  # Malignant Post
                'D': 12,  # Benign Ant
                'E': 12   # Benign Post
            }

            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Save
            wb.save(file_path)
            logging.info(f"[BSI EXPORT] Excel exported to {file_path}")
            
        except Exception as e:
            logging.info(f"[BSI EXPORT] Error: {e}")
    
    def export_chart_to_file(self, file_path: Path) -> bool: return self.bsi_canvas.export_chart(file_path)

    def export_report_to_file(self, file_path: Path) -> bool:
        if not self.current_patient_id or not self.current_study_date: return False
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write("="*60+"\nBSI QUANTIFICATION REPORT\n"+"="*60+"\n")
                f.write(f"Patient ID: {self.current_patient_id}\n")
                f.write(f"Study Date: {datetime.strptime(self.current_study_date, '%Y%m%d').strftime('%Y-%m-%d')}\n")
                f.write(f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                if hasattr(self, 'quant_manager') and self.quant_manager.current_results:
                    summary = self.quant_manager.current_results.get('summary_statistics', {})
                    mode = summary.get('processing_mode', 'unknown')
                    f.write(f"Processing Mode: {mode}\n")
                    if 'single_view' in mode: f.write(f"Note: Only {mode.split('_')[-1]} view files were available.\n")
                    f.write("\nBSI SCORES:\n" + "-"*30 + "\n")
                    if mode == 'dual_view':
                        f.write(f"Anterior BSI: {summary.get('anterior_bsi', 0.0):.10f}".rstrip('0').rstrip('.') + "\n")
                        f.write(f"Posterior BSI: {summary.get('posterior_bsi', 0.0):.10f}".rstrip('0').rstrip('.') + "\n")
                    elif mode == 'single_view_anterior':
                        f.write(f"Anterior BSI: {summary.get('anterior_bsi', 0.0):.2f}%\n")
                        f.write("Posterior BSI: N/A\n")
                    elif mode == 'single_view_posterior':
                        f.write("Anterior BSI: N/A\n")
                        f.write(f"Posterior BSI: {summary.get('posterior_bsi', 0.0):.2f}%\n")
                f.write("\nDETAILED QUANTIFICATION DATA:\n" + "-"*30 + "\n")
                headers = ["Malignant Ant", "Malignant Post", "Benign Ant", "Benign Post"]
                for r in range(self.results_table_left.rowCount()):
                    region = self.results_table_left.item(r, 0).text() if self.results_table_left.item(r, 0) else ""
                    f.write(f"{region}:\n")
                    for c, head in enumerate(headers):
                        val = self.results_table_right.item(r, c).text() if self.results_table_right.item(r, c) else "N/A"
                        f.write(f"  {head}: {val}\n")
                    f.write("\n")
                f.write("="*60+"\n")
            return True
        except Exception as e:
            logging.info(f"[BSI REPORT] Error: {e}")
            return False